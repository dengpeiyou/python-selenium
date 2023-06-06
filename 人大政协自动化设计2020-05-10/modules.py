#coding=utf-8
# 唐山人大自动登录程序 2020-03-20 by dengpeiyou
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import unittest, time, re,os,datetime,sys
import win32gui
import win32con

#用于验证码识别
import pytesseract
from PIL import Image
from PIL import ImageGrab
from PIL import ImageDraw
from collections import defaultdict

#用于读写xlsx文件
import openpyxl,gc
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

#初始化图片目录和文件名
img_path="image"
vcode_png=img_path+"/vcode.png"
vcode_2value=img_path+"/vcode2.png"
screenshot=img_path+"/save_screen.png"

dpi=1                               #<<<<<■■■■■■■■■■■■■■■■■■■■■■■■■■■■■这个是用来正确识别验证码的变量

#================================================================定义一个气泡类
class TestTaskbarIcon:
    def __init__(self):
        # 注册一个窗口类
        wc = win32gui.WNDCLASS()
        hinst = wc.hInstance = win32gui.GetModuleHandle(None)
        wc.lpszClassName = "PythonTaskbarDemo"
        wc.lpfnWndProc = {win32con.WM_DESTROY: self.OnDestroy, }
        classAtom = win32gui.RegisterClass(wc)
        style = win32con.WS_OVERLAPPED | win32con.WS_SYSMENU
        self.hwnd = win32gui.CreateWindow(classAtom, "Taskbar Demo", style,
                                          0, 0, win32con.CW_USEDEFAULT, win32con.CW_USEDEFAULT,
                                          0, 0, hinst, None)
        hicon = win32gui.LoadIcon(0, win32con.IDI_APPLICATION)
        nid = (self.hwnd, 0, win32gui.NIF_ICON, win32con.WM_USER + 20, hicon, "Demo")
        win32gui.Shell_NotifyIcon(win32gui.NIM_ADD, nid)

    def showMsg(self, title, msg):
        # 原作者使用Shell_NotifyIconA方法代替包装后的Shell_NotifyIcon方法
        # 据称是不能win32gui structure, 我稀里糊涂搞出来了.
        # 具体对比原代码.
        nid = (self.hwnd,  # 句柄
               0,  # 托盘图标ID
               win32gui.NIF_INFO,  # 标识
               0,  # 回调消息ID
               0,  # 托盘图标句柄
               "TestMessage",  # 图标字符串
               msg,  # 气球提示字符串
               0,  # 提示的显示时间
               title,  # 提示标题
               win32gui.NIIF_INFO  # 提示用到的图标
               )
        win32gui.Shell_NotifyIcon(win32gui.NIM_MODIFY, nid)

    def OnDestroy(self, hwnd, msg, wparam, lparam):
        nid = (self.hwnd, 0)
        win32gui.Shell_NotifyIcon(win32gui.NIM_DELETE, nid)
        win32gui.PostQuitMessage(0)  # Terminate the app.

#读取Excel数据
class DoExcel:
    def __init__(self,filename):
        #filename: excel文件名
        self.file = filename
        self.wk = load_workbook(self.file)

    def do_excel(self,sheetname):
        #param sheetname: 工作簿名称
        sheet = self.wk[sheetname]
        max_row = sheet.max_row #最大行
        max_column = sheet.max_column #最大列
        data = [] #定义一个空列表,用于存储所有数据
        for r in range(2,max_row+1):
            subdata = {} #定义一个字典,用于存储每行数据
            for c in range(1,max_column+1):
                key = sheet.cell(1,c).value  #取第一行表头数据
                subdata[key] = sheet.cell(r,c).value #字典格式，表头作为key
            data.append(subdata)
        return data


def 显示气泡(信息):
    tip = TestTaskbarIcon()
    tip.showMsg(信息, "^_^")
    return tip

def 关闭气泡(tip):
    win32gui.DestroyWindow(tip.hwnd)  
    return
#==================================================================================================================================1.初始化浏览器
def browser_init(url):
    user_agent = "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.116 Safari/537.36"

    opt=webdriver.ChromeOptions()
    #以下语句应用于chrome80.0
    opt.add_experimental_option("excludeSwitches", ['enable-automation'])
    opt.add_argument("disable-infobars")          #关闭安全提示条
    opt.add_argument("--start-maximized")         #启动即最大化
    opt.add_argument("--disable-popup-blocking")  #禁用弹出拦截 
    opt.add_argument("no-sandbox")                #关闭沙盘
    opt.add_argument("disable-extensions")        #扩展插件检测
    opt.add_argument("no-default-browser-check")  #默认浏览器检测

    #关闭弹出密码提示
    prefs = {"":""}
    prefs["credentials_enable_service"] = False
    prefs["profile.password_manager_enabled"] = False
    opt.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=opt)
    driver.implicitly_wait(5)
    driver.get(url)

    return driver

#==================================================================================================================================2.输入用户名和密码
def user_password_input(driver,username,password):
    IDlogin_xpath='//button[@class="el-button el-button--primary"]//span[contains(text(), "账号密码登录")]/..'
    username_xpath='//input[@placeholder="账号"]'
    password_xpath='//input[@type="password"]'

    
    登录按钮=driver.find_element_by_xpath(IDlogin_xpath)
    用户名输入框=driver.find_element_by_xpath(username_xpath)
    密码输入框  =driver.find_element_by_xpath(password_xpath)
   
    登录按钮.click()

    用户名输入框.clear()
    用户名输入框.send_keys(username)
    密码输入框.clear()
    密码输入框.send_keys(password)


#==================================================================================================================================3.屏幕截图,找到目标图红框标注并保存
def save_vcode(driver,screen_pict,vcode_xpath,vcode_pict_name):

    #保存截图
    driver.save_screenshot(screen_pict) 

    #目标控件的位置
    vcode_pict=driver.find_element_by_xpath(vcode_xpath)
    vcode_url=vcode_pict.get_attribute('src')
    left   = vcode_pict.location['x']
    top    = vcode_pict.location['y']
    right  = vcode_pict.location['x'] + vcode_pict.size['width']
    bottom = vcode_pict.location['y'] + vcode_pict.size['height']
    vcode_pos=(left*dpi, top*dpi, right*dpi, bottom*dpi) #目标控件的尺寸vcode_pos

    #截取目标控件小图
    im = Image.open(screen_pict)
    im = im.crop(vcode_pos)
    im.save(vcode_pict_name)
    
    #在大图上给目标控件画框
    im = Image.open(screen_pict)
    pict_size=im.size                    #窗口截图的尺寸pict_size
    obj=ImageDraw.ImageDraw(im)
    obj.rectangle(vcode_pos,outline="red",width=2)
    im.save(screen_pict)


#==================================================================================================================================4.获取图片中像素点数量最多的像素
def get_threshold(image):
    pixel_dict = defaultdict(int)

    # 像素及该像素出现次数的字典
    rows, cols = image.size
    for i in range(rows):
        for j in range(cols):
            pixel = image.getpixel((i, j))
            pixel_dict[pixel] += 1

    count_max = max(pixel_dict.values()) # 获取像素出现出多的次数
    pixel_dict_reverse = {v:k for k,v in pixel_dict.items()}
    threshold = pixel_dict_reverse[count_max] # 获取出现次数最多的像素点

    return threshold

#==================================================================================================================================5.按照阈值进行二值化处理
def get_bin_table(threshold):
    # 获取灰度转二值的映射table(threshold: 像素阈值)
    table = []
    for i in range(256):
        rate = 0.001 # 在threshold的适当范围内进行处理(这个阀值相当重要)
        if threshold*(1-rate)<= i <= threshold*(1+rate):
            table.append(1)
        else:
            table.append(0)
    return table

#==================================================================================================================================6.除噪
def cut_noise(image):

    rows, cols = image.size # 图片的宽度和高度
    change_pos = [] # 记录噪声点位置

    # 遍历图片中的每个点，除掉边缘
    for i in range(1, rows-1):
        for j in range(1, cols-1):
            # pixel_set用来记录该店附近的黑色像素的数量
            pixel_set = []
            # 取该点的邻域为以该点为中心的九宫格
            for m in range(i-1, i+2):
                for n in range(j-1, j+2):
                    if image.getpixel((m, n)) != 1: # 1为白色,0位黑色
                        pixel_set.append(image.getpixel((m, n)))

            # 如果该位置的九宫内的黑色数量小于等于4，则判断为噪声
            if len(pixel_set) <= 4:
                change_pos.append((i,j))

    # 对相应位置进行像素修改，将噪声处的像素置为1（白色）
    for pos in change_pos:
        image.putpixel(pos, 1)

    return image # 返回修改后的图片

#==================================================================================================================================7.灰度图、二值化去孤点
def pict_process(image):
    image = Image.open(image) # 打开图片文件
    imgry = image.convert('L')  # 转化为灰度图

    # 获取图片中的出现次数最多的像素，即为该图片的背景
    max_pixel = get_threshold(imgry)

    # 将图片进行二值化处理
    table = get_bin_table(max_pixel)
    out = imgry.point(table, '1')

    # 去掉图片中的噪声（孤立点）
    out = cut_noise(out)

    #保存图片
    out.save(vcode_2value)

#==================================================================================================================================8.清理临时截图文件
def clear_pict():
    os.remove(screenshot) 
    os.remove(vcode_png)
    os.remove(vcode_2value)
    os.rmdir(img_path)    

#==================================================================================================================================9.处理验证码
def input_vcode(driver):
    vcode_input_xpath='//input[@placeholder="验证码"]'
    vcode_id='//*[@id="VCode"]'

    if not os.path.exists(img_path):              #新建图片目录
        os.makedirs(img_path)
    save_vcode(driver,screenshot,vcode_id,vcode_png)     #截出验证码

    #识别验证码
    pict_process(vcode_png)                       #灰度图、二值化去孤点
    image = Image.open(vcode_2value)              #除噪
    text=pytesseract.image_to_string(image,lang='eng',\
        config='--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789').strip()
    clear_pict()                                  #清理临时图片(屏蔽可用于调试)

    #写入验证码(这里会找到两个符合条件的,取第2个)
    vcode_input=driver.find_elements_by_xpath(vcode_input_xpath)[1]
    vcode_input.clear()
    vcode_input.send_keys(text)

#==================================================================================================================================10.用户登录
def user_login(driver,username,password):
    user_password_input(driver,username,password)                       #输入用户名和密码
    input_vcode(driver)                                                 #处理验证码
    IDLogin_Xpath='//button[@class="el-button el-button--primary"]//span[contains(text(), "账号登录")]/..'
    user_login_buttons=driver.find_elements_by_xpath(IDLogin_Xpath)[1]  #这里取到两个符合条件的,取第二个
    user_login_buttons.click()
    time.sleep(4)
    
#==================================================================================================================================11.日期时间转特定格式字符串
def time_add(add_number):
    return datetime.datetime.now()+datetime.timedelta(minutes=add_number)

#==================================================================================================================================12.三个参数分别表示延迟多少分钟开始测试，议程持继时间，会议持续时间
def make_time(delay_time,yisheng_time,huiyi_time):
    #默认5分钟后开始测试
    now_5 = time_add(delay_time)
    str_begin=now_5.strftime('%Y-%m-%d %H:%M')



    #指定会议时间间隔
    add_minute=huiyi_time+delay_time
    str_end=(time_add(add_minute)).strftime("%Y-%m-%d %H:%M")



    #定义议程开始和结束时间
    now_add_10=time_add(delay_time+5)
    now_add_25=time_add(delay_time+yisheng_time)

    str_yicheng_begin=(now_add_10).strftime("%Y-%m-%d %H:%M")
    str_yicheng_end=(now_add_25).strftime("%Y-%m-%d %H:%M")

    #按顺序返回:会议开始时间、议程开始时间、议程结束时间,会议结束时间
    return str_begin,str_yicheng_begin,str_yicheng_end,str_end
    
#==================================================================================================================================13.用于去掉只读并填写日期输入框
def write_datetime(driver,time_name,str_time):
    time.sleep(0.5)
    driver.execute_script("arguments[0].removeAttribute('readonly')",time_name)
    time_name.send_keys(str_time)


#==================================================================================================================================14.向百度富文本框写入内容
def baidu_write(driver,write_content,index):
    baidu_input=driver.find_elements_by_xpath("//*[contains(@id,'ueditor_')]")
    if (index==1):
        ele=baidu_input[0]
    else:
        ele=baidu_input[1]
    driver.switch_to.frame(ele)
    baidu_text=driver.find_element_by_xpath("/html/body")
    baidu_text.send_keys(write_content)
    driver.switch_to.default_content()

#==================================================================================================================================15.从excel表里读取用户数据
def user_data(DataFileName,task,num):
    file_name=DataFileName
    wb=load_workbook(file_name)
    ws=wb[task]

    biaoti=[]
    shuju=[]
    #获取标题
    for cell in list(ws.rows)[0]:
        biaoti.append(cell.value)
    #获取第n行的所有数据
    for cell in list(ws.rows)[num]:
        shuju.append(cell.value)
    del wb,ws            #wb为打开的工作表
    gc.collect()         #马上内存就释放了。    
    return dict(zip(biaoti,shuju)) #制作成字典

#==================================================================================================================================16.获取用户数据
def ReadData(DataFileName,SheetName):
    wb=load_workbook(DataFileName)
    ws=wb[SheetName]
    DataRows=ws.max_row-1   #数据行

    DataList=[]
    for index in range(1,DataRows+1):
        #读入用户数据
        UserData=user_data(DataFileName,SheetName,index)
        DataList.append(UserData)
    del wb,ws            #wb为打开的工作表
    gc.collect()         #马上内存就释放了。
    return DataList
#==================================================================================================================================17.写一个人大活动
def 人大活动(driver,活动数据):
    #主菜单
    time.sleep(1)
    MainmenuXpath='//ul[@class="el-menu-demo el-menu-default el-menu--horizontal el-menu"]/li[contains(text(),"履职活动管理系统")]'
    WaitElement(driver,MainmenuXpath,20).click()
    #我的活动
    time.sleep(0.5)
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"我的活动")]').click()

    time.sleep(0.5)
    driver.find_element_by_xpath('//i[@class="el-icon-plus"]/..').click()

    #标题
    ele_title=driver.find_element_by_xpath("//label[@for='activityTitle']/../div/div[1]/input")
    ele_title.clear()
    ele_title.send_keys(活动数据['会议标题'])


    slelect_down=driver.find_elements_by_xpath("(//div[@class='el-input el-input--suffix'])")

    #类型
    slelect_down[0].click()
    time.sleep(1)
    driver.find_element_by_xpath("//div[6]/div/div/ul/li[2]").click()

    #区域
    slelect_down[1].click()
    time.sleep(1)
    driver.find_element_by_xpath("//div[7]/div/div/ul/li[1]").click()

    #调用函数获取时间元组
    t1,t2,t3,t4=make_time(5,20,30)

    #获取日期框
    select_time=driver.find_elements_by_xpath("//input[@class='ivu-input ivu-input-default ivu-input-with-suffix']")
    #开始日期
    write_datetime(driver,select_time[0],t1)
    #结束日期
    write_datetime(driver,select_time[1],t4)


    #添加议程
    time.sleep(1)
    driver.find_element_by_xpath("//button[@class='el-button el-button--primary el-button--small']/span[contains(text(), '添加活动议程')]/..").click()
    driver.find_element_by_xpath("//label[@for='agendaName']/../div/div[1]/input").clear()
    driver.find_element_by_xpath("//label[@for='agendaName']/../div/div[1]/input").send_keys(活动数据['议程名'])

    #再次获取日期框(因为这两个是动态加载)
    select_time=driver.find_elements_by_xpath("//input[@class='ivu-input ivu-input-default ivu-input-with-suffix']")
    #议程开始时间
    time.sleep(0.5)
    write_datetime(driver,select_time[2],t2)
    #议程结束时间
    time.sleep(0.5)    
    write_datetime(driver,select_time[3],t3)
    #确定按钮
    driver.find_element_by_xpath("//span/button[2]/span").click()

    #参与人
    选择参与人(driver,活动数据['参与人'])
    
    #活动内容
    txt_neirong=活动数据['活动内容']
    baidu_write(driver,txt_neirong,1)


    #活动内容
    txt_anpai=活动数据['活动安排']
    baidu_write(driver,txt_anpai,2)


    #活动地址(兄弟元素定位方法)
    input_address=driver.find_element_by_xpath("//label[@for='activityAddress']/following-sibling::div//input")
    input_address.send_keys(活动数据['活动地址'])

    #滑动到发送按钮
    send_button=driver.find_element_by_xpath("//button[@class='el-button el-button--primary']/span[contains(text(), '发起')]/..")
    driver.execute_script("arguments[0].scrollIntoView();",send_button)
    time.sleep(1)
    send_button.click()
    time.sleep(5)
    
#==================================================================================================================================18.在20秒之内反复查找一个元素
def WaitElement(driver,ElementXpath,seconds):
    for i in range(1,seconds*2):
        try:
            seek_obj=driver.find_element_by_xpath(ElementXpath)
            break
        except Exception as e:
            pass
        time.sleep(0.5)

    #必须等待
    time.sleep(2)
    return seek_obj  

#==================================================================================================================================19.写一个政协闭会提案
def 闭会提案(driver,提案数据):

    #主菜单(提案)
    driver.find_element_by_xpath('//li[@role="menuitem"][1]').click()


    #闭会提案
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"提交闭会提案")]').click()

    #标题
    time.sleep(0.5)
    driver.find_element_by_xpath("//input[@type='text']").clear()
    driver.find_element_by_xpath("//input[@type='text']").send_keys(提案数据['标题'])

    #类别
    driver.find_element_by_xpath('//input[@placeholder="请选择提案类别"]').click()
    类别列表=driver.find_elements_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"]')[2]
    time.sleep(0.5)
    i=提案数据['类别']
    类别列表.find_element_by_xpath(".//li[{}]".format(i)).click()

    baidu_write(driver,提案数据['内容'],1)


    #评分
    driver.find_element_by_xpath("(//input[@type='text'])[9]").clear()
    driver.find_element_by_xpath("(//input[@type='text'])[9]").send_keys(提案数据['评分'])

    #提交
    driver.find_element_by_xpath('//button[@class="el-button el-button--primary"]/span[contains(text(), "提交提案")]/..').click()

    #点击查看
    time.sleep(1)
    Ele_Xpath='//button[@class="el-button el-button--default el-button--small el-button--primary "]'
    ViewList=WaitElement(driver,Ele_Xpath,10)
    ViewList.send_keys(Keys.ENTER) #不知为何这里click()事件报错，只能发回车了
    time.sleep(1)
#==================================================================================================================================20.写一个社情民意
def 社情民意(driver,数据):
    #主菜单(社情民意)
    driver.find_element_by_xpath('//li[@role="menuitem"][2]').click()
    #提交社情民意
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"提交社情民意")]').click()
    
    #标题
    time.sleep(0.5)
    EleTitle=driver.find_element_by_xpath('//label[@for="socialConditionsTitle"]/following-sibling::div/div/input')
    EleTitle.clear()
    EleTitle.send_keys(数据['标题'])
    

    #类别
    driver.find_element_by_xpath('//input[@placeholder="请选择社情民意类别"]').click()
    TypeList=driver.find_element_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"][count(li)=7]')
    time.sleep(0.5)
    i=数据['类别']
    TypeList.find_element_by_xpath(".//li[{}]".format(i)).click()

    #写入内容
    baidu_write(driver,数据['内容'],1)


    #提交单位(1:研究室，2:专委会)
    driver.find_element_by_xpath('//input[@placeholder="请选择提交单位"]').click()
    e1=driver.find_element_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"][count(li)=2]')
    time.sleep(0.5)
    index=数据['提交单位']
    e1.find_element_by_xpath("./li[{}]/span".format(index)).click()


    #提交社情民意
    driver.find_element_by_xpath('//button[@class="el-button el-button--primary"]/span[contains(text(), "提交社情民意")]/..').click()
    time.sleep(1)

    #这里用于提示重复提案时确认
    try:
        driver.find_element_by_xpath('//button[@class="el-button el-button--primary"]/span[contains(text(), "确认提交")]/..').click()
    except Exception as e:
        pass
    
    #点击查看已提交
    Ele_Xpath='//button[@class="el-button el-button--default el-button--small el-button--primary "]'
    ViewList=WaitElement(driver,Ele_Xpath,10)
    ViewList.send_keys(Keys.ENTER) #不知为何这里click()事件报错，只能发回车了
    time.sleep(1)

def 审核对话框(driver):
    #填写审查内容
    CheckContent=driver.find_element_by_xpath('//textarea[@class="el-textarea__inner"]')
    CheckContent.clear()
    CheckContent.send_keys("pass")

    #确定审查
    YesCheck=driver.find_element_by_xpath('//span[@class="dialog-footer"][count(button)=2]/button[2]')
    YesCheck.click()

    #再确定(如果没有就秒用点了)
    try:
        time.sleep(1)
        driver.find_element_by_xpath('//button[@class="el-button el-button--default el-button--small el-button--primary "]').click()
    except:
        pass
#==================================================================================================================================21.提案委审核
def 提案委审核(driver,数据):
    #主菜单(提案)
    driver.find_element_by_xpath('//ul[@role="menubar"]/li[1]').click()
    #待处理提案
    driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[contains(text(), "待处理提案")]').click()

    time.sleep(1)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    find_content.send_keys(数据['标题'])
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询

    try:
        #审查按钮
        check=driver.find_element_by_xpath("//button[@title='审查']")        
        driver.execute_script("arguments[0].scrollIntoView(false);",check)
        check.click()
    except:
        return

    审核对话框(driver)

    疑似重复=False
    try:
        #如果3秒以上这个框还在,则为疑似重复提案
        time.sleep(3)
        YesCheck=driver.find_element_by_xpath('//span[@class="dialog-footer"][count(button)=2]/button[2]')
        #关闭按钮
        driver.find_elements_by_xpath('//span[@class="el-dialog__title"][contains(text(), "审查")]/../button')[0].click()
        疑似重复=True
    except:
        疑似重复=False
        pass

    if 疑似重复:
        #点疑似重复菜单
        time.sleep(1)
        driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[contains(text(), "疑似重复提案")]').click()

        #根据重复标题查找审查按钮
        RepeatTitle=driver.find_element_by_xpath('//div[@role="tablist"]/div/div/div/div/div/span[contains(text(),"{}")]'.format(数据['标题']))
        CheckButton=RepeatTitle.find_element_by_xpath("../following-sibling::div/button")
        CheckButton.click()

        审核对话框(driver)
        time.sleep(2)


#==================================================================================================================================22.提案委交办
def 提案委交办(driver,数据):
    #主菜单(提案)
    driver.find_element_by_xpath('//ul[@role="menubar"]/li[1]').click()
    #待处理提案
    driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[contains(text(), "待处理提案")]').click()

    time.sleep(0.5)
    driver.find_element_by_xpath('//ul[@class="el-menu-demo el-menu--horizontal el-menu"]/li[3]').click()
    time.sleep(0.5)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    #find_content.send_keys(数据['标题'])
    find_content.send_keys("机器自动生成")
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询

    #点击交办
    交办数据=False
    try:
        ProcessButton=driver.find_element_by_xpath("//button[@title='交办']")        
        driver.execute_script("arguments[0].scrollIntoView(false);",ProcessButton)
        ProcessButton.click()
        交办数据=True
    except:
        return
    
    if 交办数据:
        #点交办给承办单位
        time.sleep(0.5)
        DanweiRedio=driver.find_element_by_xpath('//span[@ class="el-radio__label"][contains(text(), "交办给承办单位")]')
        DanweiRedio.click()

        #选择单位
        time.sleep(0.5)
        e1=driver.find_element_by_xpath('//button[@class="el-button el-button--default el-button--small"]/span[contains(text(),"点击选择主办单位")]')
        e1.click()

        time.sleep(0.5)
        e2=driver.find_element_by_xpath('//input[@placeholder="输入关键字进行过滤"]')
        e2.clear()
        e2.send_keys("教育局")

        time.sleep(0.3)
        e3=driver.find_element_by_xpath('//span[@class="el-tree-node__label"][contains(text(),"教育局")]/preceding-sibling::label/span')
        e3.click()

        time.sleep(0.3)
        e4=driver.find_element_by_xpath('//button[@class="el-button el-button--primary"]/span[contains(text(),"确 定")]/..')
        e4.click()

        #退回时间
        time.sleep(0.5)
        #driver.find_element_by_xpath('//input[@class="el-button el-button--default el-button--small"]').click()
        driver.find_element_by_xpath('//input[@placeholder="请选择办理单位申请退回时间"]').click()
        time.sleep(0.5)
        driver.find_element_by_xpath('//div[@x-placement="bottom-start"]/div/div/ul/li[1]').click()
     

        #交办意见
        time.sleep(0.5)
        driver.find_element_by_xpath("//div[@class='el-textarea']/textarea").send_keys("请办理单位及时办理为盼")

        #答复截止时间
        ThreeeMonthRedio=driver.find_element_by_xpath('//span[@ class="el-radio__label"][contains(text(), "3个月")]')
        ThreeeMonthRedio.click()
        
        #评分
        ProposalMark=driver.find_element_by_xpath('//input[@role="spinbutton"]')
        ProposalMark.send_keys(Keys.CONTROL, 'a')
        ProposalMark.send_keys("95")

        #交办按钮
        time.sleep(1)
        #driver.find_element_by_xpath('//div[@class="el-dialog__body"]/following-sibling::div/span/button[1]').click()
        driver.find_element_by_xpath('//span[@class="dialog-footer"]//span[contains(text(),"交办")]/..').click()
        time.sleep(1)
    #if语句结束

#==================================================================================================================================23.单位办理       
def 单位办理(driver,数据):
    #主菜单
    time.sleep(1)
    driver.find_element_by_xpath('//ul[@role="menubar"]/li[contains(text(),"提案办理系统")]').click()
    #左菜单
    time.sleep(1)
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"待处理提案")]').click()

    #查询
    time.sleep(1)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    find_content.send_keys(数据['标题'])
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询

    try:
        #办理
        time.sleep(0.5)
        driver.find_elements_by_xpath('*//td/div/button/span[contains(text(),"办理")]/..')[0].click()
        #确认办理
        driver.find_element_by_xpath('//div[@class="el-dialog__body"]//following-sibling::div/span/button[2]').click()
        time.sleep(1)
    except:
        pass
#==================================================================================================================================24.单位沟通   
def 单位沟通(driver,数据):
    #主菜单
    time.sleep(1)
    driver.find_element_by_xpath('//ul[@role="menubar"]/li[contains(text(),"提案办理系统")]').click()
    #左菜单
    time.sleep(1)
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"待处理提案")]').click()

    driver.find_element_by_xpath('//div[@class="sug-disposal"]/ul/li[text()="办理中提案"]').click()

    #查询
    time.sleep(1)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    find_content.send_keys(数据['标题'])
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询
    try:
        driver.find_elements_by_xpath('*//td/div/button/span[contains(text(),"沟通")]/..')[0].click()

        time.sleep(1)
        详情=driver.find_element_by_xpath('//textarea[@class="el-textarea__inner"]')
        详情.clear()
        详情.send_keys("彼此进行了友好的沟通")

        time.sleep(1)
        driver.find_element_by_xpath('//div[@aria-label="沟通"]/div[3]/span/button[2]').click()
        time.sleep(1)
    except:
        print("没有找到需要沟通的内容")
        pass
#==================================================================================================================================25.单位答复  
def 单位答复(driver,数据):
    #主菜单
    MainmenuXpath='//ul[@role="menubar"]/li[contains(text(),"提案办理系统")]'
    WaitElement(driver,MainmenuXpath,20).click()
    
    #左菜单
    time.sleep(1)
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"待处理提案")]').click()

    driver.find_element_by_xpath('//div[@class="sug-disposal"]/ul/li[text()="办理中提案"]').click()

    #查询
    time.sleep(1)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    find_content.send_keys(数据['标题'])
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询

    #点答复
    
    try:
        time.sleep(1)
        driver.find_elements_by_xpath('*//td/div/button/span[text()="答复"]/..')[0].click()
    except:
        return
    
    #选择办理结果
    driver.find_element_by_xpath('*//input[@placeholder="请选择办理结果"]').click()
    driver.find_element_by_xpath('*//div[@x-placement="bottom-start"]//li/span[contains(text(),"A类")]').click()

    write_content="这个建议很好"

    baidu_write(driver,write_content,1)

    #联系人信息
    driver.find_element_by_xpath('*//label[@for="replyName"]/following-sibling::div/div/input').send_keys("张三")
    driver.find_element_by_xpath('*//label[@for="replyPosition"]/following-sibling::div/div/input').send_keys("局长")
    driver.find_element_by_xpath('*//label[@for="replyPhone"]/following-sibling::div/div/input').send_keys("13569885588")

    #评分
    mark=driver.find_element_by_xpath('*//input[@role="spinbutton"]')
    mark.send_keys(Keys.CONTROL+'a')
    mark.send_keys("95")

    #确定
    driver.find_element_by_xpath('*//form[@class="el-form demo-ruleForm"]/../following-sibling::div/span/button[2]').click()
    time.sleep(0.5)
    #再确认
    driver.find_element_by_xpath('*//div[@class="el-message-box__btns"]/button[2]').click()
    time.sleep(1)
#==================================================================================================================================26.人大建议(闭会)  
def 人大闭会建议(driver,数据):

    #主菜单(建议办理系统)
    driver.find_element_by_xpath('//ul[@role="menubar"]/li[contains(text(),"建议办理系统")]').click()


    #提交闭会建议
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"提交闭会建议")]').click()

    #标题
    time.sleep(0.5)
    driver.find_element_by_xpath('//input[contains(@placeholder,"请输入标题")]').clear()
    driver.find_element_by_xpath('//input[contains(@placeholder,"请输入标题")]').send_keys(数据['标题'])

    #类别
    driver.find_element_by_xpath('//input[@placeholder="请选择建议类别"]').click()
    类别列表=driver.find_elements_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"]')[2]
    time.sleep(0.5)
    i=数据['类别']
    类别列表.find_element_by_xpath(".//li[{}]".format(i)).click()

    baidu_write(driver,数据['内容'],1)


    #评分
    #driver.find_element_by_xpath("(//input[@type='text'])[9]").clear()
    #driver.find_element_by_xpath("(//input[@type='text'])[9]").send_keys(数据['评分'])

    #提交
    driver.find_element_by_xpath('//span[text()="提交建议"]').click()

    #点击查看
    time.sleep(1)
    Ele_Xpath='//button[@class="el-button el-button--default el-button--small el-button--primary "]'
    ViewList=WaitElement(driver,Ele_Xpath,10)
    ViewList.send_keys(Keys.ENTER) #不知为何这里click()事件报错，只能发回车了
    time.sleep(1)
#==================================================================================================================================27.选工委审核
def 选工委审核(driver,数据):
    #主菜单
    time.sleep(10)
    MainmenuXpath='//ul[@role="menubar"]/li[contains(text(),"建议办理系统")]'
    WaitElement(driver,MainmenuXpath,30).click()
    
    
    #待处理建议
    driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[contains(text(), "待审查建议")]').click()

    time.sleep(1)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    find_content.send_keys(数据['标题'])
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询

    try:
        #审查按钮
        check=driver.find_element_by_xpath("//button[@title='审查']")        
        driver.execute_script("arguments[0].scrollIntoView(false);",check)
        check.click()
    except:
        return

    审核对话框(driver)

    疑似重复=False
    try:
        #如果3秒以上这个框还在,则为疑似重复建议
        time.sleep(3)
        YesCheck=driver.find_element_by_xpath('//span[@class="dialog-footer"][count(button)=2]/button[2]')
        #关闭按钮
        driver.find_elements_by_xpath('//span[@class="el-dialog__title"][contains(text(), "审查")]/../button')[0].click()
        疑似重复=True
    except:
        疑似重复=False
        pass

    if 疑似重复:
        #点疑似重复菜单
        time.sleep(1)
        driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[contains(text(), "疑似重复建议")]').click()

        #根据重复标题查找审查按钮
        RepeatTitle=driver.find_element_by_xpath('//div[@role="tablist"]/div/div/div/div/div/span[contains(text(),"{}")]'.format(数据['标题']))
        CheckButton=RepeatTitle.find_element_by_xpath("../following-sibling::div/button")
        CheckButton.click()

        审核对话框(driver)
        time.sleep(2)

#===================================================================================================================================28.选工委交办
def 选工委交办(driver,数据):
    #主菜单
    time.sleep(10)
    MainmenuXpath='//ul[@role="menubar"]/li[contains(text(),"建议办理系统")]'
    WaitElement(driver,MainmenuXpath,30).click()
    
    #待交办建议
    time.sleep(0.5)
    driver.find_element_by_xpath('//aside[@class="menu-expanded leftFrame"]//span[contains(text(),"待交办建议")]').click()

    time.sleep(1)
    find_content=driver.find_element_by_xpath('//input[@placeholder="请输入标题查询.."]')
    find_content.clear()
    find_content.send_keys(数据['标题'])
    driver.find_element_by_xpath("//div[@class='el-input-group__append']/button").click()   #查询
    

    #点击交办
    交办数据=False
    try:
        ProcessButton=driver.find_element_by_xpath("//button[@title='交办']")        
        driver.execute_script("arguments[0].scrollIntoView(false);",ProcessButton)
        ProcessButton.click()
        交办数据=True
    except:
        return
    
    if 交办数据:
        #点交办给办理单位
        time.sleep(0.5)
        DanweiRedio=driver.find_element_by_xpath('//span[@ class="el-radio__label"][contains(text(), "交办给办理单位")]')
        DanweiRedio.click()

        #选择单位主办和会办
        time.sleep(0.5)
        选择单位(driver,"教育局,公安局","主办")
        time.sleep(1)
        选择单位(driver,"民政局","会办")

        #退回时间
        time.sleep(0.5)
        #driver.find_element_by_xpath('//input[@class="el-button el-button--default el-button--small"]').click()
        driver.find_element_by_xpath('//input[@placeholder="请选择办理单位申请退回时间"]').click()
        time.sleep(0.5)
        driver.find_element_by_xpath('//div[@x-placement="bottom-start"]/div/div/ul/li[1]').click()
     

        #交办意见
        time.sleep(0.5)
        driver.find_element_by_xpath("//div[@class='el-textarea']/textarea").send_keys("请办理单位及时办理为盼")

        #答复截止时间
        ThreeeMonthRedio=driver.find_element_by_xpath('//span[@ class="el-radio__label"][contains(text(), "3个月")]')
        ThreeeMonthRedio.click()
        
        #评分
        #ProposalMark=driver.find_element_by_xpath('//input[@role="spinbutton"]')
        #ProposalMark.send_keys(Keys.CONTROL, 'a')
        #ProposalMark.send_keys("95")

        #交办按钮
        time.sleep(1)
        #driver.find_element_by_xpath('//div[@class="el-dialog__body"]/following-sibling::div/span/button[1]').click()
        driver.find_element_by_xpath('//span[@class="dialog-footer"]//span[contains(text(),"交办")]/..').click()
        time.sleep(1)
    #if语句结束

#================================================================================================================================================29.选择参与人
        
def 选择参与人(driver,PersonStr):
    PersonList=PersonStr.split(",")
    #选择参与人
    driver.find_element_by_xpath('//button[@class="el-button el-button--primary"]/span[text()="选择参与人"]/..').click()
    for PerElement in PersonList:
        #输入查找
        Person=driver.find_element_by_xpath('//div[@class="filter-tree el-input el-input--small"]/input[@placeholder="输入关键字进行过滤"]')
        Person.clear()
        Person.send_keys(PerElement)

        #点复选框
        time.sleep(0.5)
        driver.find_element_by_xpath('//div[@class="el-tree-node__children"]//span[contains(text(),"{}")]/preceding-sibling::label[1]'.format(PerElement)).click()

        #添加到列表
        time.sleep(0.5)        
        driver.find_element_by_xpath('//button[@class="el-button address-first-btn el-button--primary is-circle"]').click()
    #for循环结束


    #确定
    time.sleep(1)
    driver.find_elements_by_xpath('//div[@class="el-dialog__footer"]//span[text()="确 定"]/..')[2].click()

#===================================================================================================================================30.选择单位
def 选择单位(driver,单位字符,单位类型):

    提示="点击选择"+单位类型+"单位"
    driver.find_element_by_xpath(f'//button[@class="el-button el-button--default el-button--small"]/span[text()="{提示}"]/..').click()
    单位列表=单位字符.split(",")
    type_str=单位类型+"单位"
    for 单位 in 单位列表:
        查询=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="{type_str}"]//input[@placeholder="输入关键字进行过滤"]')
        查询.send_keys(Keys.CONTROL, 'a')
        查询.send_keys(Keys.BACKSPACE)

        time.sleep(0.5)
        查询.send_keys(单位)

        driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="{type_str}"]//span[text()="{单位}"]/preceding-sibling::label/span').click()
        time.sleep(0.5)
    #for循环结束
    查询.send_keys(Keys.CONTROL, 'a')
    查询.send_keys(Keys.BACKSPACE)
    time.sleep(0.5)
    
    #确定

    driver.find_element_by_xpath(f'//div[@aria-label="{type_str}"]//span[text()="确 定"]/..').click()
    time.sleep(0.5)
#===================================================================================================================================31.检查工作台按钮是否存在
def 工作台(driver):
    rows=driver.find_elements_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr/td[7]/div')
    for ele in rows:
        nums=ele.find_elements_by_tag_name("button")
        if len(nums)==3:
            return True
    #for循环结束
    return False

#===================================================================================================================================32.添加角色
def 政协添加角色(driver,角色列表):
    time.sleep(1.5)
    主页组件=driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[text()="主页组件管理"]')
    driver.execute_script("arguments[0].scrollIntoView();",主页组件)
    主页组件.click()    
    
    
    for 角色 in 角色列表:
        #{0:{2}>10}表示用对第一个变量,用第3个变量填充10个汉字长(用于中英文混排的对齐)
        #print('角色名称:{0:{2}>10}    脚色权限:{1:>8}'.format(角色['角色名称'].strip(),角色['角色权限'],chr(12288)))

          
        
        #添加
        time.sleep(0.5)
        driver.find_element_by_xpath('//button[@class="el-button el-button--primary el-button--small"]').click()    
        
        #角色名称
        time.sleep(0.5)
        name=driver.find_element_by_xpath('//div[@class="el-dialog__body"]/form/div[1]/div/div/input')
        name.clear()
        name.send_keys(角色['角色名称'])
        
        #角色权限
        time.sleep(0.5)
        acc=driver.find_element_by_xpath('//div[@class="el-dialog__body"]/form/div[2]/div/div/input')
        acc.clear()
        acc.send_keys(角色['角色权限'])
        
        #e3=driver.find_element_by_xpath('//div[@class="el-dialog__body"]/form/div[3]/div/div/input')
        #e3.send_keys("12")    
        
        
        driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="创建组件角色"]/div[3]/span/button[2]').click()
    #for循环结束
   
#===================================================================================================================================33.为角色添加组件
def 政协添加组件(driver,角色列表):
    time.sleep(1)
    主页组件=driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[text()="主页组件管理"]')
    driver.execute_script("arguments[0].scrollIntoView();",主页组件)
    主页组件.click()  
    
    time.sleep(1)
    driver.find_element_by_xpath('//div[@class="el-input el-input--mini el-input--suffix"]/span').click()     #点击每页显示列数
    time.sleep(0.5)
    
    driver.find_element_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"]/li[last()]').click() #选最大页数
    time.sleep(1)
    
    i=0
    #倒序遍历所有角色(取出列表里的每个字典进行处理)
    #for_A循环开始
    for 角色 in  reversed(角色列表):
        #获取[编辑组件]按钮
        time.sleep(1)
        row=driver.find_element_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr[last()-{i}]')
        编辑组件=row.find_element_by_xpath('./td[3]/div/button/span[contains(text(),"编辑组件")]/..')
        角色字符=row.find_element_by_xpath('./td[1]/div').text      #取网页显示的角色
        编辑组件.click()   #点击[编辑组件]按钮
        time.sleep(0.5)
        
        #如果网页显示的角色和对应的角色记录相等(数据对应)
        if (角色字符==角色['角色名称']):
            driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]//button[@class="el-button el-button--primary"]').click()    #点击【添加】按钮

            #删除这两个字典项是为了和权限列表一一对应
            del 角色['角色名称']
            del 角色['角色权限']
            
            j=1
            #把列表里行对象字典权限不为0的项勾选(遍历每一列的字典对象)
            #for_B循环开始
            for key,value in 角色.items():
                if (value!=0):
                    time.sleep(0.2)
                    CheckBox=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="添加组件"]//table/tbody/tr[{j}]/td[1]/div/label/span/span')
                    #CheckBox=driver.find_elements_by_xpath(f'//div[@role="dialog" and @aria-label="添加组件"]//table/tbody/tr[{j}]/td[1]/div/label/span/span')[0]
                    if not CheckBox.is_displayed():         #查看元素是否可点击
                        driver.execute_script("arguments[0].scrollIntoView();",CheckBox)
                        time.sleep(2)
                    try:
                        CheckBox.click()
                        #CheckBox.send_keys(Keys.ENTER)
                    except:
                        print("复选框点击失败")
                        CheckBoxs=driver.find_elements_by_xpath(f'//div[@role="dialog" and @aria-label="添加组件"]//table/tbody/tr[{j}]/td[1]/div/label/span/span')
                        print("总共搜到{0}符合条件的复选框.".format(len(CheckBoxs)))
                j=j+1
            #for_B循环开始
            
            #【确定】
            OK_Button=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="添加组件"]/div[3]/span/button[2]')
            driver.execute_script("arguments[0].scrollIntoView();",OK_Button)
            OK_Button.click()   #此时处理完成一个角色的添加了

            #再次点击【编辑组件】按钮
            time.sleep(1)
            row=driver.find_element_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr[last()-{i}]')
            编辑组件=row.find_element_by_xpath('./td[3]/div/button/span[contains(text(),"编辑组件")]/..')
            编辑组件.click()   #点击[编辑组件]按钮
            time.sleep(1)
            i=i+1  #处理指针上移一行
            
            #计算已添加的行数
            rows=driver.find_elements_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr')
            row_count=len(rows)
            #for_C循环开始(用于修改排序)
            m=1
            for m in range(row_count):
                #获取第2列的组件类型
                time.sleep(1)
                key_str=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr[{m+1}]/td[2]/div').text
                #以这个字符串为key,取出需要的序号
                sort_code2=角色[key_str]
                SortButton=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr[{m+1}]/td[7]/div/button/span[text()="修改排序"]/..')
                SortButton.click()
                
                #输入排序号
                InputSort=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="提示"]/div/div[2]/div[2]/div[1]/input')
                InputSort.clear()
                InputSort.send_keys(sort_code2)  
                
                driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="提示"]/div/div[3]/button[2]').click() #确认
                m=m+1
            #for_C循环结束
           
            # 【确定】后处理下一个角色
            time.sleep(0.5)
            ComponentOK=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="组件管理"]/div[3]/span/button[2]')
            ComponentOK.click()
    #for_A循环结束
                

#===================================================================================================================================34.为工作台添加组件
def  添加工作台组件(driver,工作台组件,路由代码,RGB色值):    
    #左侧【主页组件管理】
    time.sleep(1.5)
    主页组件=driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[text()="主页组件管理"]')
    driver.execute_script("arguments[0].scrollIntoView();",主页组件)
    主页组件.click()        
    
    #数据行数
    rows=driver.find_elements_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr')
    row_count=len(rows)
    
    time.sleep(1)
    #逐行进行处理(正序处理)
    j=1
    for row_index in range(row_count):
        time.sleep(1.5)
        row=driver.find_element_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr[{row_index+1}]')
        编辑组件=row.find_element_by_xpath('./td[3]/div/button/span[contains(text(),"编辑组件")]/..')
        角色字符=row.find_element_by_xpath('./td[1]/div').text   #第一列(角色配置的主要依据 )

        #角色列表-->【编辑组件】点击
        编辑组件.click()   
        time.sleep(1)          

        #如果工作台【编辑组件】按钮不在，程序立刻返回
        if not 工作台(driver):
            #如果没有工作台点【取消】退出检查下一个
            ComponentManageCancel=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[3]/span/button[1]')
            ComponentManageCancel.click()   
            time.sleep(0.5)
            continue    #如果没有工作台取消后执行下一个


        #【工作台】-->【编辑组件】(查找有3个按钮的表格行，看哪个是工作台编辑)
        #组件列表-->【编辑组件】点击
        driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]//table/tbody/tr/td[7]/div[count(button)=3]/button[1]').click()
        time.sleep(0.5)
        
        j=j+1
        
        #如果有工作台(循环处理添加所有工作台组件)
        for 组件 in 工作台组件:
            if (角色字符==组件["角色名称"]):
                ColorIndex=0                   #颜色值复位
                #从符合条件的记录字典里取出所有组件项(如果值为非0的项添加上去)
                for key,value in  组件.items():                
                    if (value!=0 and key!="角色名称"):
                        #【添加】按钮
                        e1=driver.find_elements_by_xpath('//div[@role="dialog" and @aria-label="工作台组件管理"]/div[2]/button/span[contains(text(),"添 加")]/..')[0]
                        e1.click()   
                        #组件名称
                        ComponentForm=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="添加工作台组件"]/div[2]/form')
                        ComponentName=ComponentForm.find_element_by_xpath('./div[1]/div/div/input')
                        ComponentName.clear()
                        ComponentName.send_keys(key)
                        
                        
                        #政协路由代码
                        code=路由代码[key]
                        RouteCode=driver.find_element_by_xpath('//label[@class="el-form-item__label"][text()="组件路由"]/following-sibling::div/div/input')
                        RouteCode.clear()
                        RouteCode.send_keys(code)
 
                        #组件序号
                        ComponentSort=ComponentForm.find_element_by_xpath('./div[3]/div/div/input')
                        ComponentSort.clear()
                        ComponentSort.send_keys(value)
                        
                        #点击颜色选择器
                        driver.find_element_by_xpath('//label[@class="el-form-item__label"][text()="组件颜色"]/following-sibling::div/div').click()
                        
                        #颜色值输入框
                        time.sleep(0.3)
                        ColorValueInput=driver.find_element_by_xpath('//div[@class="el-color-dropdown el-color-picker__panel"]/div[2]/span/div/input')
                        ColorValueInput.clear()
                        ColorValueInput.send_keys(RGB色值[ColorIndex])   #取色值                     
                        ColorIndex= ColorIndex+1   #色值加1，下一个按钮颜色
                        
                        #颜色【确定】
                        driver.find_element_by_xpath('//button[@class="el-button el-color-dropdown__btn el-button--default el-button--mini is-plain"]').click()
                        #添加工作台组件【确定】
                        driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="添加工作台组件"]/div[3]/span/button[2]').click()
                        time.sleep(0.5)

                #for循环结束            
                        
                #工作台组件管理【确定】
                try:
                    PlatManagetOK=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="工作台组件管理"]/div[3]/span/button') 
                    PlatManagetOK.click() 
                    time.sleep(0.5)
                except:
                    pass
                #组件管理【确定】
                try:
                    driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[3]/span/button[2]').click()  
                    time.sleep(0.5)
                except:
                    pass

            else:
                print("显示角色和模板不一致，请检查模板角色名。")
#===================================================================================================================================35.人大添加角色
def 添加角色(driver,角色列表):
    
    for 角色 in 角色列表:
        #{0:{2}>10}表示用对第一个变量,用第3个变量填充10个汉字长(用于中英文混排的对齐)
        #print('角色名称:{0:{2}>10}    脚色权限:{1:>8}'.format(角色['角色名称'].strip(),角色['角色权限'],chr(12288)))
        
        #添加
        time.sleep(0.5)
        driver.find_element_by_xpath('//button[@class="el-button el-button--primary el-button--small"]').click()    
        
        #角色名称
        time.sleep(0.5)
        name=driver.find_element_by_xpath('//div[@class="el-dialog__body"]/form/div[1]/div/div/input')
        name.clear()
        name.send_keys(角色['角色名称'])
        
        #角色权限
        time.sleep(0.5)
        acc=driver.find_element_by_xpath('//div[@class="el-dialog__body"]/form/div[2]/div/div/input')
        acc.clear()
        acc.send_keys(角色['角色权限'])
        
        #e3=driver.find_element_by_xpath('//div[@class="el-dialog__body"]/form/div[3]/div/div/input')
        #e3.send_keys("12")    
        
        
        driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="创建组件角色"]/div[3]/span/button[2]').click()
    #for循环结束
   
#===================================================================================================================================36.人大角色添加组件
def 添加组件(driver,角色列表):
    time.sleep(1)
    driver.find_element_by_xpath('//div[@class="el-input el-input--mini el-input--suffix"]/span').click()     #点击每页显示列数
    time.sleep(0.5)
    
    driver.find_element_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"]/li[last()]').click() #选最大页数
    time.sleep(1)
    
    i=0
    #倒序遍历所有角色(取出列表里的每个字典进行处理)
    #for_A循环开始
    for 角色 in  reversed(角色列表):
        #获取[编辑组件]按钮
        time.sleep(1)
        row=driver.find_element_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr[last()-{i}]')
        编辑组件=row.find_element_by_xpath('./td[3]/div/button/span[contains(text(),"编辑组件")]/..')
        角色字符=row.find_element_by_xpath('./td[1]/div').text      #取网页显示的角色
        编辑组件.click()   #点击[编辑组件]按钮
        time.sleep(0.5)
        
        #如果网页显示的角色和对应的角色记录相等(数据对应)
        if (角色字符==角色['角色名称']):
            driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]//button[@class="el-button el-button--primary"]').click()    #点击【添加】按钮

            #删除这两个字典项是为了和权限列表一一对应
            del 角色['角色名称']
            del 角色['角色权限']
            
            j=1
            #把列表里行对象字典权限不为0的项勾选(遍历每一列的字典对象)
            #for_B循环开始
            for key,value in 角色.items():
                if (value!=0):
                    time.sleep(0.2)
                    CheckBox=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="添加组件"]//table/tbody/tr[{j}]/td[1]/div/label/span/span')
                    #CheckBox=driver.find_elements_by_xpath(f'//div[@role="dialog" and @aria-label="添加组件"]//table/tbody/tr[{j}]/td[1]/div/label/span/span')[0]
                    if not CheckBox.is_displayed():         #查看元素是否可点击
                        driver.execute_script("arguments[0].scrollIntoView();",CheckBox)
                        time.sleep(2)
                    try:
                        CheckBox.click()
                        #CheckBox.send_keys(Keys.ENTER)
                    except:
                        print("复选框点击失败")
                        CheckBoxs=driver.find_elements_by_xpath(f'//div[@role="dialog" and @aria-label="添加组件"]//table/tbody/tr[{j}]/td[1]/div/label/span/span')
                        print("总共搜到{0}符合条件的复选框.".format(len(CheckBoxs)))
                j=j+1
            #for_B循环开始
            
            #【确定】
            OK_Button=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="添加组件"]/div[3]/span/button[2]')
            driver.execute_script("arguments[0].scrollIntoView();",OK_Button)
            OK_Button.click()   #此时处理完成一个角色的添加了

            #再次点击【编辑组件】按钮
            time.sleep(1)
            row=driver.find_element_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr[last()-{i}]')
            编辑组件=row.find_element_by_xpath('./td[3]/div/button/span[contains(text(),"编辑组件")]/..')
            编辑组件.click()   #点击[编辑组件]按钮
            time.sleep(1)
            i=i+1  #处理指针上移一行
            
            #计算已添加的行数
            rows=driver.find_elements_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr')
            row_count=len(rows)
            #for_C循环开始(用于修改排序)
            for m in range(row_count):
                #获取第2列的组件类型
                key_str=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr[{m+1}]/td[2]/div').text
                #以这个字符串为key,取出需要的序号
                sort_code2=角色[key_str]
                
                #修改排序
                time.sleep(0.5)
                SortButton=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="组件管理"]/div[2]/div/div[3]/table/tbody/tr[{m+1}]/td[7]/div/button/span[text()="修改排序"]/..')
                SortButton.click()
                
                #输入排序号
                InputSort=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="提示"]/div/div[2]/div[2]/div[1]/input')
                InputSort.clear()
                InputSort.send_keys(sort_code2)  
                
                SortOK=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="提示"]/div/div[3]/button[2]')
                SortOK.click()
            #for_C循环结束
            #修改完一个角色之后点击确定提交
            time.sleep(0.5)
            ComponentOK=driver.find_element_by_xpath(f'//div[@role="dialog" and @aria-label="组件管理"]/div[3]/span/button[2]')
            ComponentOK.click()
    #for_A循环结束

#===================================================================================================================================37.人大超管添加租户        
def 添加租户(driver,数据):
    租户按钮=driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[text()="租户管理"]')
    driver.execute_script("arguments[0].scrollIntoView();",租户按钮)
    租户按钮.click()  
    time.sleep(1)
    
    添加=driver.find_element_by_xpath('//div[@class="grid-content bg-purple-dark"]/button/span[text()="添加"]/..')
    添加.click()
    time.sleep(1)    
    
    #填写头5个
    添加窗口=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="添加租户"]')
    
    str_list=['租户名字','显示文字','登陆地址', '登录账号','密码']
    for i in range(len(str_list)):
        tmep_text=数据[0][str_list[i]]
        添加窗口.find_element_by_xpath(f'./div[2]/form/div[{i+1}]/div/div/input').send_keys(tmep_text)
        time.sleep(0.2)
    
    #所属专委会
    添加窗口.find_element_by_xpath(f'./div[2]/form/div[10]/div/div/input').send_keys(数据[0]['所属专委会'])
    
    #选择区域
    区域字符=数据[0]['区域']
    driver.find_element_by_xpath("//div[@id='app']//form/div[9]/div/div/div/span/span/i").click()
    time.sleep(0.2)
    driver.find_element_by_xpath(f"//div[3]/div/div/ul/li/span[text()='{区域字符}']").click()
    
    #租户类型选择
    time.sleep(0.5)
    类型字符_str2=数据[0]['租户类型']
    driver.find_element_by_xpath("//div[@id='app']//form/div[11]/div/div/div/span/span/i").click()
    time.sleep(0.2)
    driver.find_element_by_xpath(f"//div[4]/div/div/ul/li/span[text()='{类型字符_str2}']").click()    
    
    
    #权重
    driver.find_element_by_xpath("//div[@id='app']//form/div[16]/div/div/span").click()
    
    #添加
    driver.find_element_by_xpath("//div[@id='app']//form/div[18]/div/button/span").click()

#===================================================================================================================================38.政协删除工作台组件   
def  删除工作台组件(driver,工作台组件):    
    #左侧【主页组件管理】
    time.sleep(1.5)
    主页组件=driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[text()="主页组件管理"]')
    driver.execute_script("arguments[0].scrollIntoView();",主页组件)
    主页组件.click()        
    
    #数据行数
    rows=driver.find_elements_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr')
    row_count=len(rows)
    
    time.sleep(1)
    #逐行进行处理(正序处理)
    j=1
    for row_index in range(row_count):
        time.sleep(1.5)
        row=driver.find_element_by_xpath(f'//div[@class="el-table__body-wrapper is-scrolling-none"]/table/tbody/tr[{row_index+1}]')
        编辑组件=row.find_element_by_xpath('./td[3]/div/button/span[contains(text(),"编辑组件")]/..')

        #角色列表-->【编辑组件】点击
        编辑组件.click()   
        time.sleep(1)          

        #如果工作台【编辑组件】按钮不在，程序立刻返回
        if not 工作台(driver):
            #如果没有工作台点【取消】退出检查下一个
            ComponentManageCancel=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[3]/span/button[1]')
            ComponentManageCancel.click()   
            time.sleep(0.5)
            continue    #如果没有工作台取消后执行下一个


        #【工作台】-->【编辑组件】(查找有3个按钮的表格行，看哪个是工作台编辑)
        #组件列表-->【编辑组件】点击
        driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]//table/tbody/tr/td[7]/div[count(button)=3]/button[1]').click()
        time.sleep(0.5)
        
        j=j+1
        
        #如果有工作台(循环处理添加所有工作台组件)
        trs=driver.find_elements_by_xpath('//div[@role="dialog" and @aria-label="工作台组件管理"]//tbody/tr[@class="el-table__row"]') 
        for k in range(len(trs)):
            删除按钮=trs[0].find_element_by_xpath('//td[6]/div/button[2]') 
            删除按钮.click()
            time.sleep(0.3)
        #for循环结束            
                        
        #工作台组件管理【确定】
        try:
            PlatManagetOK=driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="工作台组件管理"]/div[3]/span/button') 
            PlatManagetOK.click() 
            time.sleep(0.5)
        except:
            pass
        #组件管理【确定】
        try:
            driver.find_element_by_xpath('//div[@role="dialog" and @aria-label="组件管理"]/div[3]/span/button[2]').click()  
            time.sleep(0.5)
        except:
            pass

#===================================================================================================================================39.删除角色
def 删除角色(driver,角色列表):
    time.sleep(1)
    主页组件=driver.find_element_by_xpath('//ul[@class="el-menu-vertical-demo left_menu hideBar el-menu"]/li/span[text()="主页组件管理"]')
    driver.execute_script("arguments[0].scrollIntoView();",主页组件)
    主页组件.click()  
    
    time.sleep(1)
    driver.find_element_by_xpath('//div[@class="el-input el-input--mini el-input--suffix"]/span').click()     #点击每页显示列数
    time.sleep(0.5)
    
    driver.find_element_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"]/li[last()]').click() #选最大页数
    time.sleep(1)
    
    
    i=0
    #倒序遍历所有角色(取出列表里的每个字典进行处理)
    #for_A循环开始
    time.sleep(0.5)
    
    tres=driver.find_elements_by_xpath('//table[@class="el-table__body"]/tbody/tr')
    first_row=driver.find_elements_by_xpath('//div[@class="el-table__body-wrapper is-scrolling-none"]//tr')[0]
    for i in  range(len(tres)):
        time.sleep(0.8)
        #获取【删除】按钮
        删除按钮=first_row.find_element_by_xpath('./td[3]/div/button[3]')
        删除按钮.click()   #点击[删除按钮]
        time.sleep(0.5)
        #点击提示里的【确定】按钮
        driver.find_element_by_xpath('//div[@class="el-message-box__btns"]/button[2]').click()   
        i=i+1  #处理指针上移一行
    #for_A循环结束
#===================================================================================================================================40. 添加用户管理员
def 添加用户管理员(driver,数据):
    #管理员=['测试人大','yhgly']
    
    tip=显示气泡("任务处理中...")
    #角色管理
    driver.find_element_by_xpath("//div[@id='app']/section/div/div/aside/ul/li[2]/span").click()
    time.sleep(1.5)
    #每页显示40条
    driver.find_element_by_xpath("//div[@id='app']/section/div/div/section/div/div[2]/div/div[3]/div/div/span[2]/div/div/span/span/i").click()
    time.sleep(1)
    driver.find_element_by_xpath("//div/ul/li[5]/span").click()
    time.sleep(1)
    #查找租户是否存在，不存在则添加
    租户列表=driver.find_elements_by_xpath("//div[@id='app']//table/tbody/tr/td[2]/div")
    存在=False
    for i in   range(len(租户列表)):
        if 租户列表[i].text==(数据[0]['管理员']):
            存在=True
            print("================找到了，老大!!!")    
    #添加
    if not 存在:
        driver.find_element_by_xpath("//div[@id='app']/section/div/div/section/div/div[2]/div/div/div/div/button/span").click()
        driver.find_element_by_xpath("(//input[@type='text'])[2]").clear()
        driver.find_element_by_xpath("(//input[@type='text'])[2]").send_keys(数据[0]['管理员'])
        driver.find_element_by_xpath("(//input[@type='text'])[3]").clear()
        driver.find_element_by_xpath("(//input[@type='text'])[3]").send_keys(数据[0]['角色类型'])
        time.sleep(0.5)
        driver.find_element_by_xpath("//div[@id='app']//form/div[4]/div/button/span").click()
    else:
        print("你要添加的用户已经存在!")
  
    tip.showMsg("任务处理成!!!", " ")

#===================================================================================================================================41. 读取管理员资源文件
def 读取资源(FileName,SheetName):
    workbook1=load_workbook(FileName)
    sheet=workbook1[SheetName]
    max_row=sheet.max_row
    datas=[]
    for i in range(2,max_row+1):
        row_data=[]
        row_data.append(sheet.cell(i,2).value)
        row_data.append(sheet.cell(i,3).value)
        datas.append(row_data)
    del workbook1,sheet  #workbook1为打开的工作表
    gc.collect()         #马上内存就释放了。         
    return datas